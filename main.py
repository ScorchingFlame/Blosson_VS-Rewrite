import json
from typing import Iterator
username = input("Username: ")
password = input("Password: ")

with open("./config.json", "r") as f:
    da = json.loads(f.read())
    if username != da["admin"]["USERNAME"] or password != da["admin"]["PASSWORD"]:
        exit()
from ctypes import WinError
from threading import Thread
from time import sleep
from flask_compress import Compress
from flask_socketio import SocketIO
from flask_session import Session
from concurrent.futures import ThreadPoolExecutor
from io import StringIO
from tempfile import NamedTemporaryFile
import flask, sqlite3, socket, random, os, re, base64, re, logging, atexit, run
from flask import abort, redirect, request, session, url_for
from flask import Flask, render_template
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader


# Additional Imports For EXE                 
from engineio.async_drivers import gevent

def get_ip_address():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.connect(("8.8.8.8", 80))
    return s.getsockname()[0]

def change_date_format(dt):
        return re.sub(r'(\d{4})-(\d{1,2})-(\d{1,2})', '\\3-\\2-\\1', dt)

def clear():
  
    # for windows
    if os.name == 'nt':
        _ = os.system('cls')
  
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = os.system('clear')
  
if not os.path.exists("./static/pics"):
    os.makedirs("./static/pics")
network = True
ip = ""
try:
    ip = get_ip_address()
except OSError:
    network = False
    # print("No Network Found, Pls connect To A Network To Work For Other Devices, At This Time You cant access This App In Other Devices")
app = Flask(__name__)
socketio = SocketIO(app,async_mode='gevent')
compress = Compress()
compress.init_app(app)
executor = ThreadPoolExecutor(2)
# executer = Executor(app)
conn = sqlite3.connect('data/voting.db', check_same_thread=False)
conn.execute("""CREATE TABLE IF NOT EXISTS positions (
	            id INTEGER PRIMARY KEY,
	            name TEXT NOT NULL,
	            wcs TEXT NOT NULL
                );""")
conn.execute("""CREATE TABLE IF NOT EXISTS candidates ( 
	            id INTEGER PRIMARY KEY,
	            name TEXT NOT NULL,
	            STD TEXT NOT NULL,
                House TEXT NOT NULL,
                Photo TEXT NOT NULL,
                Position INTEGER NOT NULL,
                Votes INTEGER NOT NULL
                );""")
conn.execute("""CREATE TABLE IF NOT EXISTS voters (
                adnumber INTEGER NOT NULL,
                name TEXT NOT NULL,
                STD TEXT NOT NULL,
                House TEXT NOT NULL,
                Voted INTEGER NOT NULL
                );""")
with open('config.json', 'r') as cfg:
    CFG = json.loads(cfg.read())
    if CFG['debug'] == False:
        app.logger.disabled = True
        log = logging.getLogger('werkzeug')
        log.disabled = True
        os.environ['WERKZEUG_RUN_MAIN'] = 'true'
    else:
        log = StringIO()
        logger = logging.getLogger('werkzeug')
        stream = logging.StreamHandler(log)
        os.environ['WERKZEUG_RUN_MAIN'] = 'true'
        logger.addHandler(stream)

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

class FileHandler():

    def __init__(self):
        self.file = NamedTemporaryFile(suffix='.xlsx', delete=False)
        # register function called when quit
        atexit.register(self._cleanup)

    def write_into(self, btext):
        self.file.write(btext)
        self.file.seek(0)

    def _cleanup(self):
        # because self.file has been created without delete=False, closing the file causes its deletion 
        self.file.close()
        os.unlink(self.file.name)


def decode_base64(data, altchars=b'+/'):
    """Decode base64, padding being optional.

    :param data: Base64 data as an ASCII byte string
    :returns: The decoded byte string.

    """
    data = re.sub(rb'[^a-zA-Z0-9%s]+' % altchars, b'', data)  # normalize
    missing_padding = len(data) % 4
    if missing_padding:
        data += b'='* (4 - missing_padding)
    return base64.b64decode(data, altchars)

@app.route('/admin')
def index():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template('index.html')

@app.route('/admin/positions')
def positions():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template('positions.html', tm=conn.cursor().execute('SELECT * from positions').fetchall())

@app.route('/admin/positions/create', methods=['GET', 'POST'])
def create():
    if not session.get("login"):
        return redirect("/admin/login")
    if request.method == 'GET':
        return render_template('create-position.html', ph="", mode_name="Submit", bck_btn="../positions")
    if request.method == 'POST':
        name = flask.request.values.get('name') # Your form's
        wcs_str = flask.request.values.get('wcs')
        conn.execute(f"INSERT INTO positions (name, wcs) VALUES ('{name}', '{wcs_str}')")
        conn.commit()   
        return redirect('/admin/positions')

@app.route('/admin/positions/edit/<id>', methods=['GET','POST'])
def edit_positions(id: int):
    if not session.get("login"):
        return redirect("/admin/login")
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from positions where id={str(id)}")
    record = cursor.fetchone()
    if record is None:
        return redirect('/admin/positions')
    if request.method == 'GET':
        return render_template('create-position.html', ph= record[1], mode_name="Edit", bck_btn="../../positions")
    if request.method == 'POST':
        cursor = conn.cursor()
        cursor.execute(f"SELECT * from positions where id={str(id)}")
        record = cursor.fetchall()
        name = flask.request.values.get('name') # Your form's
        wcs_str = flask.request.values.get('wcs')
        cursor = conn.cursor()
        cursor.execute(f"UPDATE positions SET name='{name}', wcs='{wcs_str}' WHERE id = {id}")
        conn.commit()
        return redirect('/admin/positions')

@app.route('/admin/positions/delete/<id>')
def delete_positions(id):
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from positions where id={str(id)}")
    record = cursor.fetchone()
    if record is None:
        return redirect('/admin/positions')
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from positions where id != {id}")
    record2 = cursor.fetchall()
    rec3 = cursor.execute(f"SELECT * FROM candidates where Position != {id}")
    new_record = []
    new_r_c = []
    for o in record2:
        pl = list(o)
        pl[0] = None
        al = tuple(pl)
        new_record.append(al)
    for o in rec3:
        pl = list(o)
        pl[0] = None
        al = tuple(pl)
        new_r_c.append(al)
    sql_statement = 'INSERT INTO positions VALUES (?, ?, ?)'
    sql_statement_c = 'INSERT INTO candidates VALUES (?, ?, ?, ?, ?, ?, ?)'
    cur = conn.cursor()
    cur.execute("DELETE FROM positions")
    cur.execute("DELETE FROM candidates")
    cur.executemany(sql_statement, new_record)
    cur.executemany(sql_statement_c, new_r_c)
    conn.commit()
    return redirect('/admin/positions')

@socketio.on('llol')
def lmfo(data):
    print(data)

@app.route('/admin/candidates')
def candidates():
    if not session.get("login"):
        return redirect("/admin/login")
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from candidates ORDER BY name")
    record = cursor.fetchall()
    cursor.execute(f"SELECT * FROM positions")
    recordpos = cursor.fetchall()
    new_r = []
    for i in recordpos:
        for u,j in enumerate(record):
            if i[0] == j[5]:
                j = list(j)
                j[5] = i[1]
                # print(tuple(j), u)
                new_r.append(tuple(j))
                # record.pop(u)
    return render_template('candidates.html', tm=new_r)

@app.route('/admin/candidates/create', methods=['GET', 'POST'])
def createc():
    if not session.get("login"):
        return redirect("/admin/login")
    if request.method == 'GET':
        return render_template('create-candidates.html', ph="", phs=0, mode_name="Create", bck_btn="../candidates", poses=conn.execute('SELECT * from positions').fetchall())
    if request.method == 'POST':
        fola = request.files['photo']
        ffname = f"{''.join(random.choice('0123456789ABCDEFMJWksiwl') for i in range(6))}.{fola.filename.split('.')[-1]}"
        fola.save(f"./static/pics/{ffname}")
        name = flask.request.values.get('name') # Your form's
        std = flask.request.values.get('std')
        house = flask.request.values.get('house')
        pos = flask.request.values.get('pos')
        votes = flask.request.values.get('votes')
        conn.execute(f"INSERT INTO candidates (name, STD, House, Photo, Position, Votes) VALUES ('{name}', '{std}', '{house}', '{ffname}', {pos}, {votes})")
        conn.commit()   
        return redirect('/admin/candidates')


@app.route('/admin/candidates/delete/<id>')
def delete_candidate(id):
    if not session.get("login"):
        return redirect("/admin/login")
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from candidates where id={str(id)}")
    record = cursor.fetchone()
    if record is None:
        return redirect('/admin/candidates')
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from candidates where id != {str(id)}")
    record2 = cursor.fetchall()
    os.remove(f'./static/pics/{record[4]}')
    new_record = []
    for o in record2:
        pl = list(o)
        pl[0] = None
        al = tuple(pl)
        new_record.append(al)
    sql_statement = 'INSERT INTO candidates VALUES (?, ?, ?, ?, ?, ?, ?)'
    cur = conn.cursor()
    cur.execute("DELETE FROM candidates")
    cur.executemany(sql_statement, new_record)
    conn.commit()
    return redirect('/admin/candidates')


@app.route('/admin/candidates/edit/<id>', methods=['GET','POST'])
def edit_candidates(id: int):
    if not session.get("login"):
        return redirect("/admin/login")
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from candidates where id={str(id)}")
    record = cursor.fetchone()
    if record is None:
        return redirect('/admin/candidates')
    if request.method == 'GET':
        return render_template('create-candidates.html', ph= record[1], phs=record[2], phv=record[6] , mode_name="Edit", bck_btn="../../candidates", poses=conn.execute('SELECT * from positions').fetchall())
    if request.method == 'POST':
        cursor = conn.cursor()
        cursor.execute(f"SELECT * from candidates where id={str(id)}")
        record = cursor.fetchone()
        try:
            os.remove(f'./static/pics/{record[4]}')
        except:
            pass
        fola = request.files['photo']
        ffname = f"{''.join(random.choice('0123456789ABCDEFMJWksiwl') for i in range(6))}.{fola.filename.split('.')[-1]}"
        fola.save(f"./static/pics/{ffname}")
        name = flask.request.values.get('name') # Your form's
        std = flask.request.values.get('std')
        house = flask.request.values.get('house')
        pos = flask.request.values.get('pos')
        votes = flask.request.values.get('votes')
        cursor = conn.cursor()
        cursor.execute(f"UPDATE candidates SET name='{name}', STD='{std}', House='{house}', Photo='{ffname}', Position={pos}, Votes={votes} WHERE id = {id}")
        conn.commit()
        return redirect('/admin/candidates')

@app.route('/admin/voters')
def voters():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template('voters.html', rec=conn.cursor().execute('SELECT * from voters').fetchall())

class devnull:
    write = lambda _: None

@app.route('/admin/voters/create', methods=['GET', 'POST'])
def vcreate():
    if not session.get("login"):
        return redirect("/admin/login")
    if request.method == 'GET':
        return render_template('create-voters-manual.html')
    if request.method == 'POST':
        ad_num = request.values.get('ad')
        name = request.values.get('name')
        std = request.values.get('std')
        house = request.values.get('house')
        voted = request.values.get('voted')
        cur = conn.cursor()
        cur.execute(f"INSERT INTO voters (adnumber, name, STD, House, Voted) VALUES ({str(ad_num)}, '{name}', '{std}', '{house}', {str(voted)})")
        conn.commit()
        return redirect('/admin/voters')

@app.route('/admin/voters/delete/<ad>')
def delv(ad ):
    if not session.get("login"):
        return redirect("/admin/login")
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from voters where adnumber={str(ad)}")
    record = cursor.fetchone()
    if record is None:
        return redirect('/admin/voters')
    cursor = conn.cursor()
    cursor.execute(f"SELECT * from voters where adnumber != {str(ad)}")
    record2 = cursor.fetchall()
    cursor.execute("DELETE FROM voters")
    cursor.executemany("INSERT INTO voters VALUES (?,?,?,?,?)", record2)
    conn.commit()
    return redirect('/admin/voters')


@app.route('/admin/voters/excel')
def vexcel():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template('create-voters-excel.html')

@app.route('/upload/vexcel', methods=['POST'])
def uexcel():
    file = request.files['file'].read()
    try:
        fh = FileHandler()
        fh.write_into(file)
            # print(fh.file.name)
            # sleep(60)
        loc = (fh.file.name)
        a = load_workbook(loc)
        sheet = a.worksheets[0]
        # sheet.cell(0,0)
        # print([x.value for x in list(sheet.rows)[0]])
        if not [x.value for x in list(sheet.rows)[0]] == ["AdmissionNumber", "Name", "STD", "House"]:
                # socketio.emit('sleeep', {'done': 'ok'})
            return abort(500)

            # k = [int, str, int, Literal["WINTER", "SUMMER", "SPRING"]]
            # sleep(60)
        tp = []
        for i in range(2, sheet.max_row+1):
            meh = [str(i) if x == 1 or x == 3 else int(i) for x, i in enumerate([x.value for x in list(sheet.rows)[i-1]])]
            meh.append(0)
            tp.append(tuple(meh))

        cur = conn.cursor()
        cur.executemany("INSERT INTO voters VALUES (?, ?, ?, ?, ?)", tp)
        conn.commit()
        cur.close()
        return "Ok"
    except Exception as e:
        # print(e)
        socketio.emit('excel-error', {'error': str(e)})
        return abort(500)

@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        name = request.form.get('username')
        passwd = request.form.get('password')
        if CFG['admin']['USERNAME'] == name and CFG['admin']['PASSWORD'] == passwd:
            session["login"] = "Yea"
            return redirect('/admin')
        else:
            return render_template('login.html', error="Error: Login Credentials Are Wrong!")
    return render_template('login.html')

@app.route('/admin/cleardata/positions')
def cdpos():
    if not session.get("login"):
        return redirect("/admin/login")
    conn.cursor().execute("DELETE from positions;").execute("DELETE from candidates")
    dir = './static/pics'
    for f in os.listdir(dir):
        os.remove(os.path.join(dir, f))
    conn.commit()
    return redirect('/admin')

@app.route('/admin/cleardata/candidates')
def cdcad():
    if not session.get("login"):
        return redirect("/admin/login")
    conn.cursor().execute("DELETE from candidates")
    dir = './static/pics'
    for f in os.listdir(dir):
        os.remove(os.path.join(dir, f))
    conn.commit()
    return redirect('/admin')

@app.route('/admin/cleardata/voters')
def cdvot():
    if not session.get("login"):
        return redirect("/admin/login")
    conn.cursor().execute("DELETE from voters")
    conn.commit()
    return redirect('/admin')


@app.route('/admin/cleardata/all')
def cdall():
    if not session.get("login"):
        return redirect("/admin/login")
    conn.cursor().execute("DELETE from voters;").execute(" Delete from candidates;").execute("delete from positions")
    dir = './static/pics'
    for f in os.listdir(dir):
        os.remove(os.path.join(dir, f))
    conn.commit()
    return redirect('/admin')

@app.route('/')
def voting():
    return render_template('voting.html')

@socketio.on('req-info')
def rifo(data):
    damn = conn.cursor().execute(f"select * from voters where adnumber={data['adnumber']} and house='{data['house']}'").fetchone()
    socketio.emit('send-info', {'info': damn}, room=request.sid)

@socketio.on('connect')
def on_connect():
    rec_cad = conn.cursor().execute('SELECT * FROM candidates ORDER BY name').fetchall()
    rec_pos = conn.cursor().execute('SELECT * FROM positions').fetchall()
    rec = {'cad': rec_cad,
            'pos': rec_pos
            }
    socketio.emit('cad-with-pos', rec,room=request.sid)

@socketio.on('voted')
def voted(data):
    try:
        cursor = conn.cursor()
        # print(data)
        for x in data['voting_data']:
            cursor.execute(f"UPDATE candidates SET Votes = Votes + 1 WHERE id = {data['voting_data'][x]}")
        # print(data)
        cursor.execute(f"UPDATE voters SET Voted = 1 WHERE adnumber = {data['voter_data'][0]} and House = '{data['voter_data'][3]}'")  
        conn.commit()  
        socketio.emit('voted-complete', {'voted': True}, room=request.sid)
        socketio.emit('live-feed-data', {"voter_data": data['voter_data']})
    except Exception as e:
        socketio.emit('voted-complete', {'voted': False, 'error': str(e)}, room=request.sid)

@app.route('/admin/live-feed')
def live_feed():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template('live-feed.html')

@socketio.on('refresh')
def refresh(data):
    rec_cad = conn.cursor().execute('SELECT * FROM candidates').fetchall()
    rec_pos = conn.cursor().execute('SELECT * FROM positions').fetchall()
    new_rec = []
    for i in rec_pos:
        k = {}
        k[i[1]] = sorted([e for e in rec_cad if e[5] == i[0]], key=lambda x:x[6], reverse=True)
        new_rec.append(k)
    # print(new_rec)
    socketio.emit('income-refresh', new_rec, room=request.sid)

@app.route('/admin/logout')
def logout():
    if not session.get("login"):
        return redirect("/admin/login")
    session.clear()
    return redirect("/admin/login")

@app.route('/admin/candidates/excel')
def excel_cad():
    if not session.get("login"):
        return redirect("/admin/login")
    return render_template("create-candidates-excel.html")

@app.route('/upload/cexcel', methods=['POST'])
def upload_cexcel():
    file = request.files['file'].read()
    try:
        fh = FileHandler()
        fh.write_into(file)
            # print(fh.file.name)
            # sleep(60)
        loc = (fh.file.name)
        wb = load_workbook(loc)
        sheet = wb.worksheets[0]
        image_loader = SheetImageLoader(sheet)
        # Put your sheet in the loader
        if not [x.value for x in list(sheet.rows)[0]] == ["Name", "STD", "House" ,"Position", "Image", "StartingVotes"]:
            abort(500)
        caddo = []
        indeximage  =2
        for i in range(2, sheet.max_row+1):
            imagename = f"{''.join(random.choice('0123456789ABCDEFMJWksiwl') for i in range(9))}.jpeg"
            image_loader.get(f"E{str(indeximage)}").save(f"./static/pics/{imagename}")
            meh = [str(i) if x == 0 or x == 2 else (imagename if x == 4 else int(i)) for x, i in enumerate([x.value for x in list(sheet.rows)[i-1]])]
            indeximage+=1
            caddo.append(tuple(meh))

        cur = conn.cursor()
        cur.executemany("INSERT INTO candidates (name, STD, House, Position , Photo, Votes) VALUES (?, ?, ?, ?,?,?)", caddo)
        conn.commit()
        cur.close()
        return "Ok"
    except Exception as e:
        # print(e)
        socketio.emit('excel-c-error', {'error': str(e)})
        return abort(500)
if __name__ == '__main__':
    # server = Thread(target=lambda:app.run(host=CFG['HOST'], port=CFG['PORT']))
    # server.start()
    # asyncio.get_event_loop().create_task()
    # asyncio.get_event_loop().run_in_executor(None, lambda:run.run(network, ip, CFG['PORT'] ,CFG['admin']['USERNAME'], CFG['admin']['PASSWORD']))
    # sleep(10)
    try:
        run.run(network, ip, CFG['PORT'] , CFG['admin']['USERNAME'], CFG['admin']['PASSWORD'])
        socketio.run(app, host=CFG['HOST'], port=CFG['PORT'])
    except KeyboardInterrupt:
        pass
    except OSError:
        clear()
        print(
"""There is Another Voting System is running, So This Application Cant Be Run At the same Time
Try These: 
    * Try Closing The Other Voting System If There is no Voting System running except this, Try Restarting the System
    * Changing PORT in config.json (Contact Ritheesh)"""
)
        input("Press Enter To Exit")
    # asyncio.get_event_loop().run_in_executor(None, lambda:socketio.run(app, host=CFG['HOST'], port=CFG['PORT']))
    # server = Thread(target=lambda:socketio.run(app, host=CFG['HOST'], port=CFG['PORT']))
    # server.start()
    
    # print(log.getvalue())
    
    # server.join(5)
# http_server = WSGIServer(('0.0.0.0', 8080), app, handler_class=WebSocketServer) 
# asyncio.get_event_loop().run_in_executor(None, http_server.serve_forever)
# http_server.serve_forever()
# time.sleep(50)
# http_server.stop()